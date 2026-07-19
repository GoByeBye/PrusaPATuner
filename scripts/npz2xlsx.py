"""Convert a captured run NPZ into an xlsx for inspection in Excel.

Each stream (loadcell, pos_x, pos_y, pos_z) becomes its own sheet with:
  * t_monotonic   -- raw monotonic-clock seconds (the analyser's reference)
  * t_rel_s       -- zero-based seconds since the first sample (Excel-chart-friendly)
  * t_wall_utc    -- ISO-8601 wall-clock UTC (when the NPZ carries an anchor)
  * <stream name> -- the value column

A `meta` sheet lists the scalar plan parameters (k_values, cycle_period_s, ...).
"""
import argparse
from datetime import UTC, datetime

import numpy as np
from openpyxl import Workbook

# Map of (timestamp-array-name, value-array-name, sheet-name).
STREAMS = [
    ("force_t", "force_y", "loadcell"),
    ("pos_t",   "pos_x",   "pos_x"),
    ("pos_y_t", "pos_y",   "pos_y"),
    ("pos_z_t", "pos_z",   "pos_z"),
]
SCALAR_KEYS = [
    "sweep_t0", "k_values", "cycle_period_s", "cycles_per_K",
    "slow_half_s", "fast_half_s", "slow_feed_mm_s", "fast_feed_mm_s",
    "mono_anchor_mono", "mono_anchor_unix", "started_at_unix",
]


def _wall_clock_anchor(d) -> tuple[float, float] | None:
    """Return (mono_anchor_mono, mono_anchor_unix) if both are in the NPZ.

    With these two scalars, any monotonic timestamp `m` converts to a
    unix timestamp via `m - mono_anchor_mono + mono_anchor_unix`.
    """
    if "mono_anchor_mono" in d.files and "mono_anchor_unix" in d.files:
        try:
            return float(d["mono_anchor_mono"][0]), float(d["mono_anchor_unix"][0])
        except (IndexError, ValueError, TypeError):
            return None
    return None


def npz2xlsx(src: str, dst: str) -> None:
    with np.load(src) as d:
        anchor = _wall_clock_anchor(d)
        workbook = Workbook()
        workbook.remove(workbook.active)

        for tk, vk, sheet in STREAMS:
            if tk not in d.files or vk not in d.files:
                continue
            t = np.ravel(d[tk])
            v = np.ravel(d[vk])
            if t.size == 0 or v.size == 0:
                continue
            if t.size != v.size:
                raise ValueError(f"{tk} and {vk} have different lengths")

            worksheet = workbook.create_sheet(sheet)
            header = ["t_monotonic", "t_rel_s"]
            if anchor is not None:
                header.append("t_wall_utc")
            header.append(sheet)
            worksheet.append(header)

            for timestamp, value in zip(t, v, strict=True):
                row = [float(timestamp), float(timestamp - t[0])]
                if anchor is not None:
                    mono_a, unix_a = anchor
                    wall_time = float(timestamp) - mono_a + unix_a
                    row.append(
                        datetime.fromtimestamp(wall_time, UTC)
                        .isoformat()
                        .replace("+00:00", "Z")
                    )
                row.append(value.item() if isinstance(value, np.generic) else value)
                worksheet.append(row)

        # meta sheet: one column per scalar
        meta_cols: dict[str, list[object]] = {}
        for k in SCALAR_KEYS:
            if k in d.files:
                arr = np.ravel(d[k])
                meta_cols[k] = [v.item() if isinstance(v, np.generic) else v for v in arr]
                if not meta_cols[k]:
                    meta_cols[k] = [None]
        if meta_cols:
            worksheet = workbook.create_sheet("meta")
            worksheet.append(list(meta_cols))
            max_len = max(len(v) for v in meta_cols.values())
            for row_index in range(max_len):
                worksheet.append(
                    [
                        values[row_index] if row_index < len(values) else None
                        for values in meta_cols.values()
                    ]
                )

        if not workbook.sheetnames:
            workbook.create_sheet("meta")
        workbook.save(dst)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert NPZ run dump to XLSX")
    parser.add_argument("src", help="source .npz file")
    parser.add_argument("dst", nargs="?", help="destination .xlsx file")
    args = parser.parse_args()
    dst = args.dst or args.src.rsplit(".", 1)[0] + ".xlsx"
    npz2xlsx(args.src, dst)
    print(f"wrote {dst}")
