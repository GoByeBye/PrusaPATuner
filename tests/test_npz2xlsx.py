from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from scripts.npz2xlsx import npz2xlsx


def test_npz2xlsx_uses_declared_dependencies_and_preserves_streams(tmp_path: Path):
    source = tmp_path / "run.npz"
    destination = tmp_path / "run.xlsx"
    np.savez(
        source,
        force_t=np.array([10.0, 10.5]),
        force_y=np.array([1.25, 2.5]),
        mono_anchor_mono=np.array([10.0]),
        mono_anchor_unix=np.array([1_700_000_000.0]),
        k_values=np.array([0.02, 0.04]),
        cycles_per_K=np.array([3]),
    )

    npz2xlsx(str(source), str(destination))

    workbook = load_workbook(destination, read_only=True, data_only=True)
    assert workbook.sheetnames == ["loadcell", "meta"]
    assert list(workbook["loadcell"].values) == [
        ("t_monotonic", "t_rel_s", "t_wall_utc", "loadcell"),
        (10, 0, "2023-11-14T22:13:20Z", 1.25),
        (10.5, 0.5, "2023-11-14T22:13:20.500000Z", 2.5),
    ]
    assert list(workbook["meta"].values) == [
        ("k_values", "cycles_per_K", "mono_anchor_mono", "mono_anchor_unix"),
        (0.02, 3, 10, 1_700_000_000),
        (0.04, None, None, None),
    ]
